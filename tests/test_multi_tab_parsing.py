"""
Unit tests for multi-tab Google Sheets parsing functionality.

Tests the ability to parse multiple tabs including Prog-rock tabs.
"""

import pytest
from unittest.mock import Mock, patch, MagicMock
from io import BytesIO

from catalog.services.google_sheets import (
    GoogleSheetsService,
    is_prog_metal_tab,
    normalize_tab_name,
    extract_year,
)


class TestTabFiltering:
    """Tests for tab filtering functions."""

    def test_is_prog_metal_tab_accepts_prog_metal(self):
        """Test that Prog-metal tabs are accepted."""
        assert is_prog_metal_tab("2025 Prog-metal") is True
        assert is_prog_metal_tab("2024 Prog-metal") is True

    def test_is_prog_metal_tab_accepts_prog_rock(self):
        """Test that Prog-rock tabs are accepted."""
        assert is_prog_metal_tab("2025 Prog-rock") is True
        assert is_prog_metal_tab("2024 Prog-rock") is True

    def test_is_prog_metal_tab_accepts_year_only(self):
        """Test that year-only tabs are accepted."""
        assert is_prog_metal_tab("2025") is True
        assert is_prog_metal_tab("2017") is True

    def test_is_prog_metal_tab_rejects_other_tabs(self):
        """Test that non-music tabs are rejected."""
        assert is_prog_metal_tab("Statistics") is False
        assert is_prog_metal_tab("2025 Reissues") is False
        assert is_prog_metal_tab("Info") is False
        assert is_prog_metal_tab("") is False


class TestMultiTabParsing:
    """Tests for multi-tab parsing functionality."""

    @pytest.fixture
    def test_xlsx_path(self):
        """Return path to test XLSX file with real data."""
        from pathlib import Path
        return Path(__file__).parent / 'contract' / 'testdata' / 'progmetal_releases_2025.xlsx'

    @pytest.fixture
    def real_workbook(self, test_xlsx_path):
        """Load the real test workbook with Prog-rock tabs containing Atomiste."""
        from openpyxl import load_workbook
        return load_workbook(test_xlsx_path)

    def test_enumerate_tabs_finds_prog_rock(self, real_workbook):
        """Test that enumerate_tabs finds Prog-rock tabs in real test data."""
        service = GoogleSheetsService("https://example.com/test.xlsx")
        tabs = service.enumerate_tabs(real_workbook)

        # Should find multiple tabs including Prog-rock tabs
        assert len(tabs) > 0

        # Find the 2025 Prog-rock tab
        prog_rock_tab = next((t for t in tabs if t.name == "2025 Prog-rock"), None)
        assert prog_rock_tab is not None, f"2025 Prog-rock tab not found. Available tabs: {[t.name for t in tabs]}"
        assert prog_rock_tab.is_prog_metal is True
        assert prog_rock_tab.year == 2025

    def test_filter_tabs_includes_prog_rock(self, real_workbook):
        """Test that filter_tabs includes Prog-rock tabs from real test data."""
        service = GoogleSheetsService("https://example.com/test.xlsx")
        all_tabs = service.enumerate_tabs(real_workbook)
        filtered_tabs = service.filter_tabs(all_tabs)

        # Should have multiple tabs including both Prog-rock and Prog-metal
        tab_names = [t.name for t in filtered_tabs]
        assert "2025 Prog-rock" in tab_names, f"2025 Prog-rock not in filtered tabs: {tab_names}"
        assert "2025 Prog-metal" in tab_names, f"2025 Prog-metal not in filtered tabs: {tab_names}"
        assert "Statistics" not in tab_names, f"Statistics should be filtered out but found in: {tab_names}"

    def test_fetch_albums_from_prog_rock_tab(self, real_workbook):
        """Test fetching albums from real 2025 Prog-rock tab."""
        service = GoogleSheetsService("https://example.com/test.xlsx")
        albums = service.fetch_albums_from_tab(real_workbook, "2025 Prog-rock")

        # Should have many albums (real data has hundreds)
        assert len(albums) > 0, "No albums found in 2025 Prog-rock tab"

        # Verify Atomiste album is present
        atomiste_albums = [album for album in albums if album["artist"] == "Atomiste"]
        assert len(atomiste_albums) > 0, (
            f"Album by Atomiste not found in 2025 Prog-rock tab. "
            f"Total albums: {len(albums)}, Sample artists: {[a['artist'] for a in albums[:10]]}"
        )

        # Verify the first Atomiste album has required fields
        atomiste_album = atomiste_albums[0]
        assert "album" in atomiste_album
        assert atomiste_album["album"], "Album name is empty"
        assert "spotify_url" in atomiste_album
        assert atomiste_album["spotify_url"].startswith("https://open.spotify.com/album/")

        # Verify Atomiste has the correct genre from Google Sheets
        # Note: Raw data from sheet is "Experimental Big Band", but this gets
        # mapped to "Progressive Metal" in the database by _map_genre()
        assert "genre" in atomiste_album, "Genre field missing from Atomiste album"
        assert atomiste_album["genre"] == "Experimental Big Band", (
            f"Expected raw genre from sheet to be 'Experimental Big Band', "
            f"got: '{atomiste_album.get('genre', 'N/A')}'. "
            f"Note: This will be mapped to 'Progressive Metal' when imported to database."
        )

    def test_prog_rock_tab_artist_atomiste_exists(self, real_workbook):
        """Test that 2025 Prog-rock tab contains an album by artist Atomiste using real test data."""
        service = GoogleSheetsService("https://example.com/test.xlsx")

        # Enumerate and filter tabs
        all_tabs = service.enumerate_tabs(real_workbook)
        prog_metal_tabs = service.filter_tabs(all_tabs)

        # Find 2025 Prog-rock tab
        prog_rock_tab = next(
            (t for t in prog_metal_tabs if t.name == "2025 Prog-rock"), None
        )
        assert prog_rock_tab is not None, (
            f"2025 Prog-rock tab not found. Available tabs: {[t.name for t in prog_metal_tabs]}"
        )

        # Fetch albums from Prog-rock tab
        albums = service.fetch_albums_from_tab(real_workbook, prog_rock_tab.name)

        # Check that Atomiste exists
        atomiste_albums = [a for a in albums if a["artist"] == "Atomiste"]
        assert len(atomiste_albums) > 0, (
            f"No albums by Atomiste found in {prog_rock_tab.name} tab. "
            f"Total albums in tab: {len(albums)}. "
            f"Sample artists: {[a['artist'] for a in albums[:20]]}"
        )

        # Verify the album has required fields
        atomiste_album = atomiste_albums[0]
        assert "album" in atomiste_album
        assert atomiste_album["album"], "Atomiste album name is empty"
        assert "spotify_url" in atomiste_album
        assert atomiste_album["spotify_url"].startswith("https://open.spotify.com/album/"), (
            f"Invalid Spotify URL: {atomiste_album['spotify_url']}"
        )

        # Verify Atomiste has the correct genre from Google Sheets
        # Note: Raw data from sheet is "Experimental Big Band", but this gets
        # mapped to "Progressive Metal" in the database by _map_genre()
        assert "genre" in atomiste_album, "Genre field missing from Atomiste album"
        assert atomiste_album["genre"] == "Experimental Big Band", (
            f"Expected raw genre from sheet to be 'Experimental Big Band', "
            f"got: '{atomiste_album.get('genre', 'N/A')}'. "
            f"Note: This will be mapped to 'Progressive Metal' when imported to database."
        )

        print(f"\n✓ Found Atomiste album: '{atomiste_album['album']}' in {prog_rock_tab.name}")
        if len(atomiste_albums) > 1:
            print(f"  Total Atomiste albums in tab: {len(atomiste_albums)}")

    def test_sort_tabs_orders_prog_rock_and_metal(self, real_workbook):
        """Test that both Prog-rock and Prog-metal tabs are sorted chronologically using real test data."""
        service = GoogleSheetsService("https://example.com/test.xlsx")

        all_tabs = service.enumerate_tabs(real_workbook)
        filtered_tabs = service.filter_tabs(all_tabs)
        sorted_tabs = service.sort_tabs_chronologically(filtered_tabs)

        # Verify tabs are sorted by year
        years = [t.year for t in sorted_tabs if t.year is not None]
        assert years == sorted(years), f"Tabs not sorted chronologically: {years}"

        # Verify we have both 2024 and 2025 tabs for both types
        tab_names = [t.name for t in sorted_tabs]

        # Check that 2024 tabs (if present) come before 2025 tabs
        if "2024 Prog-rock" in tab_names and "2025 Prog-rock" in tab_names:
            idx_2024_rock = tab_names.index("2024 Prog-rock")
            idx_2025_rock = tab_names.index("2025 Prog-rock")
            assert idx_2024_rock < idx_2025_rock, "2024 Prog-rock should come before 2025 Prog-rock"

        if "2024 Prog-metal" in tab_names and "2025 Prog-metal" in tab_names:
            idx_2024_metal = tab_names.index("2024 Prog-metal")
            idx_2025_metal = tab_names.index("2025 Prog-metal")
            assert idx_2024_metal < idx_2025_metal, "2024 Prog-metal should come before 2025 Prog-metal"


class TestRealGoogleSheetsProgRock:
    """
    Integration test with real Google Sheets URL.

    This test connects to the actual Google Sheets document to verify
    that the 2025 Prog-rock tab exists and contains Atomiste.

    Requires GOOGLE_SHEETS_XLSX_URL environment variable.
    """

    @pytest.mark.skip(reason="Integration test - requires real Google Sheets access. Run manually with: pytest tests/test_multi_tab_parsing.py::TestRealGoogleSheetsProgRock -v")
    def test_real_google_sheets_prog_rock_atomiste(self):
        """Test parsing real Google Sheets for Prog-rock tab with Atomiste."""
        import os
        import requests
        from openpyxl import load_workbook

        sheets_url = os.getenv("GOOGLE_SHEETS_XLSX_URL")
        if not sheets_url:
            pytest.skip("GOOGLE_SHEETS_XLSX_URL not configured")

        # Fetch the actual workbook
        try:
            response = requests.get(sheets_url, timeout=30)
            response.raise_for_status()
            workbook = load_workbook(BytesIO(response.content))
        except Exception as e:
            pytest.skip(f"Could not fetch Google Sheets: {e}")

        service = GoogleSheetsService(sheets_url)

        # Enumerate tabs
        all_tabs = service.enumerate_tabs(workbook)
        tab_names = [t.name for t in all_tabs]

        # Check if 2025 Prog-rock exists
        prog_rock_tabs = [t for t in all_tabs if "Prog-rock" in t.name and "2025" in t.name]
        if not prog_rock_tabs:
            pytest.skip(f"2025 Prog-rock tab not found. Available tabs: {tab_names}")

        prog_rock_tab = prog_rock_tabs[0]

        # Fetch albums from Prog-rock tab
        albums = service.fetch_albums_from_tab(workbook, prog_rock_tab.name)

        # Look for Atomiste
        atomiste_albums = [a for a in albums if "Atomiste" in a["artist"]]

        assert len(atomiste_albums) > 0, (
            f"No albums by Atomiste found in {prog_rock_tab.name}. "
            f"Total albums in tab: {len(albums)}. "
            f"Sample artists: {[a['artist'] for a in albums[:10]]}"
        )

        print(f"\n✓ Found {len(atomiste_albums)} album(s) by Atomiste in {prog_rock_tab.name}")
        for album in atomiste_albums:
            print(f"  - {album['album']} ({album.get('release_date', 'unknown date')})")
