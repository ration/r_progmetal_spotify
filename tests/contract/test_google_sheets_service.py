"""
Tests for Google Sheets XLSX fetcher service.
"""

import os
from pathlib import Path
import pytest
from unittest.mock import Mock, patch

from catalog.services.google_sheets import GoogleSheetsService


@pytest.fixture
def test_xlsx_path():
    """Return path to test XLSX file."""
    return Path(__file__).parent / 'testdata' / 'progmetal_releases_2025.xlsx'


@pytest.fixture
def mock_sheets_service(test_xlsx_path):
    """Create GoogleSheetsService with mocked HTTP request."""
    service = GoogleSheetsService("https://example.com/test.xlsx")

    # Mock the fetch to return local file content
    with open(test_xlsx_path, 'rb') as f:
        test_content = f.read()

    with patch('catalog.services.google_sheets.requests.get') as mock_get:
        mock_response = Mock()
        mock_response.content = test_content
        mock_response.raise_for_status = Mock()
        mock_get.return_value = mock_response

        yield service


class TestGoogleSheetsService:
    """Tests for GoogleSheetsService class."""

    def test_initialization(self):
        """Test service initialization with URL."""
        url = "https://docs.google.com/spreadsheets/d/ABC123/export?format=xlsx&gid=123"
        service = GoogleSheetsService(url)

        assert service.xlsx_url == url

    def test_extract_url_from_hyperlink(self, mock_sheets_service, test_xlsx_path):
        """Test extracting URL from cell with hyperlink object."""
        from openpyxl import load_workbook

        wb = load_workbook(test_xlsx_path)
        ws = wb.active

        # Find first cell with hyperlink (row 7, col 9 based on our verification)
        cell = ws.cell(row=7, column=9)

        url = mock_sheets_service._extract_url_from_cell(cell)

        assert url is not None
        assert url.startswith('https://open.spotify.com/album/')
        assert len(url) > 40  # Should have album ID

    def test_extract_url_from_formula(self, mock_sheets_service, test_xlsx_path):
        """Test extracting URL from HYPERLINK formula."""
        from openpyxl import load_workbook

        wb = load_workbook(test_xlsx_path)
        ws = wb.active

        # Find first cell with HYPERLINK formula (row 8, col 9 based on verification)
        cell = ws.cell(row=8, column=9)

        url = mock_sheets_service._extract_url_from_cell(cell)

        assert url is not None
        assert url.startswith('https://open.spotify.com/album/')

    def test_extract_url_from_empty_cell(self, mock_sheets_service):
        """Test extracting URL from empty cell returns None."""
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        empty_cell = ws.cell(row=1, column=1)

        url = mock_sheets_service._extract_url_from_cell(empty_cell)

        assert url is None

    def test_find_header_row(self, mock_sheets_service, test_xlsx_path):
        """Test finding header row in worksheet."""
        from openpyxl import load_workbook

        wb = load_workbook(test_xlsx_path)
        ws = wb.active

        header_row = mock_sheets_service._find_header_row(ws)

        assert header_row == 6  # Based on actual test data

    def test_find_header_row_not_found(self, mock_sheets_service):
        """Test finding header row raises ValueError when not found."""
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active

        with pytest.raises(ValueError, match="Could not find header row"):
            mock_sheets_service._find_header_row(ws)

    def test_fetch_albums(self, mock_sheets_service):
        """Test fetching albums from XLSX file."""
        albums = mock_sheets_service.fetch_albums()

        # Verify we got albums
        assert len(albums) > 0
        assert len(albums) > 2000  # Should have 2000+ albums

        # Verify first album structure
        first_album = albums[0]
        assert 'artist' in first_album
        assert 'album' in first_album
        assert 'spotify_url' in first_album
        assert 'release_date' in first_album
        assert 'genre' in first_album
        assert 'vocal_style' in first_album
        assert 'country' in first_album

        # Verify data types
        assert isinstance(first_album['artist'], str)
        assert isinstance(first_album['album'], str)
        assert isinstance(first_album['spotify_url'], str)

        # Verify Spotify URL format
        assert first_album['spotify_url'].startswith('https://open.spotify.com/album/')

    def test_fetch_albums_filters_empty_rows(self, mock_sheets_service):
        """Test that fetch_albums filters out rows without artist/album."""
        albums = mock_sheets_service.fetch_albums()

        # All albums should have artist and album names
        for album in albums:
            assert album['artist']
            assert album['album']
            assert len(album['artist']) > 0
            assert len(album['album']) > 0

    def test_fetch_albums_filters_missing_spotify(self, mock_sheets_service):
        """Test that fetch_albums filters out rows without Spotify URL."""
        albums = mock_sheets_service.fetch_albums()

        # All albums should have Spotify URL
        for album in albums:
            assert album['spotify_url']
            assert album['spotify_url'].startswith('https://open.spotify.com/album/')

    def test_parse_release_date_full(self, mock_sheets_service):
        """Test parsing full date (YYYY-MM-DD format from Excel)."""
        from datetime import date

        # Excel dates come through as date objects already
        test_date = date(2025, 1, 15)
        result = mock_sheets_service.parse_release_date("January 15", 2025)

        assert result == test_date

    def test_parse_release_date_month_only(self, mock_sheets_service):
        """Test parsing month-only date."""
        from datetime import date

        result = mock_sheets_service.parse_release_date("January", 2025)

        assert result == date(2025, 1, 1)

    def test_parse_release_date_invalid(self, mock_sheets_service):
        """Test parsing invalid date returns None."""
        result = mock_sheets_service.parse_release_date("invalid date", 2025)

        assert result is None

    def test_parse_release_date_empty(self, mock_sheets_service):
        """Test parsing empty date returns None."""
        result = mock_sheets_service.parse_release_date("", 2025)

        assert result is None

    def test_expected_columns_present(self, mock_sheets_service):
        """Test that expected columns are present in fetched data."""
        albums = mock_sheets_service.fetch_albums()

        if albums:
            first_album = albums[0]
            expected_keys = {
                'artist', 'album', 'release_date', 'genre',
                'vocal_style', 'country', 'spotify_url'
            }

            assert set(first_album.keys()) == expected_keys

    def test_data_normalization(self, mock_sheets_service):
        """Test that data is properly normalized (stripped whitespace)."""
        albums = mock_sheets_service.fetch_albums()

        for album in albums[:10]:  # Check first 10
            # No leading/trailing whitespace
            assert album['artist'] == album['artist'].strip()
            assert album['album'] == album['album'].strip()

            # URLs should not have whitespace
            assert ' ' not in album['spotify_url']
