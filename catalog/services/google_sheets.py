"""
Google Sheets XLSX fetcher service.

This module provides functionality to fetch XLSX data from Google Sheets export URLs.
Unlike CSV exports, XLSX format preserves hyperlinks, allowing us to extract Spotify URLs.
"""

import logging
import requests
import re
from typing import List, Dict, Optional
from io import BytesIO
from datetime import datetime
from dataclasses import dataclass
from openpyxl import load_workbook

logger = logging.getLogger(__name__)


class TabProcessingError(Exception):
    """
    Tab-specific error that can be recovered from.

    Raised when a single tab fails to process but the sync can continue
    with remaining tabs.

    Examples:
        - Invalid data format in tab
        - Missing expected columns in tab
        - Empty tab with no data rows
    """

    pass


class CriticalSyncError(Exception):
    """
    System-level error that cannot be recovered from.

    Raised when the entire sync operation must abort.

    Examples:
        - Cannot reach Google Sheets or Spotify APIs
        - Disk/file I/O failure
        - System resource exhaustion
    """

    pass


@dataclass
class TabMetadata:
    """
    Metadata for a single Google Sheets tab.

    Ephemeral object used during multi-tab synchronization to filter,
    sort, and track progress across tabs.

    Attributes:
        name: Original tab name from Google Sheets
        normalized_name: Whitespace-trimmed, validated version
        year: Extracted year (e.g., 2025 from "2025 Prog-metal")
        order: Original position in workbook (0-indexed)
        is_prog_metal: True if tab passes Prog-metal filter
        estimated_rows: Optional row count from initial scan
    """

    name: str
    normalized_name: str
    year: Optional[int]
    order: int
    is_prog_metal: bool
    estimated_rows: Optional[int] = None

    def __str__(self) -> str:
        """Return string representation of TabMetadata."""
        year_str = f" ({self.year})" if self.year else ""
        return f"TabMetadata('{self.name}'{year_str}, order={self.order})"


def normalize_tab_name(raw_name: str) -> tuple[str, bool]:
    """
    Normalize tab name and validate it.

    Args:
        raw_name: Raw tab name from openpyxl

    Returns:
        Tuple of (normalized_name, is_valid)
        - normalized_name: Tab name with whitespace normalized
        - is_valid: True if tab passes validation

    Examples:
        >>> normalize_tab_name("  2025 Prog-metal  ")
        ("2025 Prog-metal", True)
        >>> normalize_tab_name("  ")
        ("", False)
    """
    if not raw_name or not isinstance(raw_name, str):
        return "", False

    # Strip leading/trailing whitespace
    normalized = raw_name.strip()

    # Reject empty after strip
    if not normalized:
        return "", False

    # Reject if contains control characters or unicode
    if not normalized.isascii():
        logger.warning(f"Tab name contains non-ASCII characters: {raw_name}")
        return "", False

    # Reject if contains newlines/tabs/control chars
    if any(c.isspace() and c not in ' ' for c in normalized):
        logger.warning(f"Tab name contains control characters: {raw_name}")
        return "", False

    # Accept names with ASCII letters, digits, spaces, dashes, underscores
    if not all(c.isalnum() or c in ' -_' for c in normalized):
        logger.warning(f"Tab name contains invalid characters: {raw_name}")
        return "", False

    return normalized, True


def extract_year(tab_name: str) -> Optional[int]:
    """
    Extract year from tab name.

    Supports:
    - "2025 Prog-metal" → 2025
    - "2024" → 2024
    - "2023x" → 2023 (fallback)

    Args:
        tab_name: Tab name from sheet

    Returns:
        Year as integer, or None if cannot extract

    Examples:
        >>> extract_year("2025 Prog-metal")
        2025
        >>> extract_year("2017")
        2017
        >>> extract_year("Statistics")
        None
    """
    if not tab_name or not isinstance(tab_name, str):
        return None

    # Pass 1: Modern format "YYYY Prog-metal"
    match = re.match(r"^(\d{4})\s+Prog-metal$", tab_name)
    if match:
        return int(match.group(1))

    # Pass 2: Legacy format "YYYY" (exactly 4 digits)
    if re.match(r"^\d{4}$", tab_name):
        return int(tab_name)

    # Pass 3: Fallback - any 4 leading digits
    match = re.match(r"^(\d{4})", tab_name)
    if match:
        return int(match.group(1))

    # No year found
    return None


def is_prog_metal_tab(tab_name: str) -> bool:
    """
    Check if tab should be imported (progressive metal or rock tab).

    Matches:
    - "2025 Prog-metal", "2024 Prog-metal" (format: YYYY Prog-metal)
    - "2025 Prog-rock", "2024 Prog-rock" (format: YYYY Prog-rock)
    - "2017", "2018", "2025" (format: YYYY only)

    Rejects:
    - "Statistics", "2025 Reissues", "Info", etc.

    Args:
        tab_name: Tab name from sheet

    Returns:
        True if tab matches progressive metal/rock pattern

    Examples:
        >>> is_prog_metal_tab("2025 Prog-metal")
        True
        >>> is_prog_metal_tab("2025 Prog-rock")
        True
        >>> is_prog_metal_tab("2017")
        True
        >>> is_prog_metal_tab("Statistics")
        False
    """
    # Rule 1: Ends with " Prog-metal" (case-sensitive)
    if tab_name.endswith(" Prog-metal"):
        return True

    # Rule 2: Ends with " Prog-rock" (case-sensitive)
    if tab_name.endswith(" Prog-rock"):
        return True

    # Rule 3: Exactly 4 digits (year format)
    if re.match(r"^\d{4}$", tab_name):
        return True

    return False


class GoogleSheetsService:
    """
    Service for fetching XLSX data from Google Sheets export URLs.

    XLSX format preserves hyperlinks, allowing extraction of Spotify URLs.

    Attributes:
        xlsx_url: The Google Sheets XLSX export URL
    """

    # Expected column names from the r/progmetal releases sheet
    EXPECTED_COLUMNS = {
        "Artist",
        "Album",
        "Release Date",
        "Length",
        "Genre / Subgenres",
        "Vocal Style",
        "Country / State",
        "Spotify",
    }

    def __init__(self, xlsx_url: str):
        """
        Initialize the Google Sheets service.

        Args:
            xlsx_url: Google Sheets XLSX export URL (must be publicly accessible)
        """
        self.xlsx_url = xlsx_url
        logger.info(f"Initialized GoogleSheetsService with URL: {xlsx_url}")

    def enumerate_tabs(self, workbook) -> List[TabMetadata]:
        """
        Enumerate all tabs in the workbook with metadata.

        Args:
            workbook: openpyxl workbook object

        Returns:
            List of TabMetadata objects for all tabs in the workbook

        Example:
            >>> workbook = load_workbook(xlsx_file)
            >>> service = GoogleSheetsService(url)
            >>> tabs = service.enumerate_tabs(workbook)
            >>> [tab.name for tab in tabs]
            ['2025 Prog-metal', '2024 Prog-metal', 'Statistics']
        """
        tabs = []
        for order, sheet_name in enumerate(workbook.sheetnames):
            # Normalize and validate tab name
            normalized, is_valid = normalize_tab_name(sheet_name)

            if not is_valid:
                logger.warning(f"Skipping invalid tab name: {sheet_name}")
                continue

            # Extract year
            year = extract_year(normalized)

            # Check if prog-metal tab
            is_pm = is_prog_metal_tab(normalized)

            # Get estimated row count
            sheet = workbook[sheet_name]
            estimated_rows = sheet.max_row if sheet.max_row else 0

            tab_metadata = TabMetadata(
                name=sheet_name,
                normalized_name=normalized,
                year=year,
                order=order,
                is_prog_metal=is_pm,
                estimated_rows=estimated_rows,
            )

            tabs.append(tab_metadata)
            logger.debug(f"Enumerated tab: {tab_metadata}")

        logger.info(f"Enumerated {len(tabs)} tabs from workbook")
        return tabs

    def filter_tabs(self, tabs: List[TabMetadata]) -> List[TabMetadata]:
        """
        Filter tabs to only include progressive metal tabs.

        Args:
            tabs: List of TabMetadata objects

        Returns:
            Filtered list containing only prog-metal tabs

        Example:
            >>> all_tabs = service.enumerate_tabs(workbook)
            >>> metal_tabs = service.filter_tabs(all_tabs)
            >>> [tab.name for tab in metal_tabs]
            ['2025 Prog-metal', '2024 Prog-metal']
        """
        filtered = [tab for tab in tabs if tab.is_prog_metal]

        logger.info(
            f"Filtered {len(filtered)} prog-metal tabs out of {len(tabs)} total tabs"
        )

        for tab in tabs:
            if not tab.is_prog_metal:
                logger.debug(f"Skipped non-prog-metal tab: {tab.name}")

        return filtered

    def sort_tabs_chronologically(self, tabs: List[TabMetadata]) -> List[TabMetadata]:
        """
        Sort tabs chronologically by extracted year (oldest to newest).

        Tabs without extractable year are appended at the end in original order.

        Args:
            tabs: List of TabMetadata objects

        Returns:
            Sorted list (oldest year first, newest year last)

        Example:
            >>> tabs = service.enumerate_tabs(workbook)
            >>> sorted_tabs = service.sort_tabs_chronologically(tabs)
            >>> [tab.name for tab in sorted_tabs]
            ['2023 Prog-metal', '2024 Prog-metal', '2025 Prog-metal']
        """
        # Separate into two groups
        with_year = [t for t in tabs if t.year is not None]
        without_year = [t for t in tabs if t.year is None]

        # Sort "with year" group numerically ascending (oldest first)
        with_year.sort(key=lambda t: t.year)

        # Log any tabs without year
        for tab in without_year:
            logger.warning(
                f"Tab '{tab.name}' (order {tab.order}) does not have extractable year. "
                f"Will be processed after dated tabs."
            )

        # Combine and return
        sorted_tabs = with_year + without_year

        if sorted_tabs:
            logger.info(
                f"Sorted {len(sorted_tabs)} tabs chronologically: "
                f"{', '.join(t.name for t in sorted_tabs[:5])}"
                f"{' ...' if len(sorted_tabs) > 5 else ''}"
            )

        return sorted_tabs

    def fetch_albums_from_tab(
        self, workbook, tab_name: str, tab_year: Optional[int] = None
    ) -> List[Dict[str, str]]:
        """
        Fetch albums from a specific tab in the workbook.

        Args:
            workbook: openpyxl workbook object
            tab_name: Name of the tab to fetch from
            tab_year: Year extracted from tab name (e.g., 2025 from "2025 Prog-metal")

        Returns:
            List of album dictionaries (same format as fetch_albums())
            Each dictionary includes 'tab_year' field if tab_year was provided

        Raises:
            KeyError: If tab_name does not exist in workbook
            ValueError: If tab has invalid structure

        Example:
            >>> workbook = load_workbook(xlsx_file)
            >>> service = GoogleSheetsService(url)
            >>> albums = service.fetch_albums_from_tab(workbook, "2025 Prog-metal", 2025)
            >>> albums[0]['tab_year']
            2025
        """
        if tab_name not in workbook.sheetnames:
            raise KeyError(f"Tab '{tab_name}' not found in workbook")

        worksheet = workbook[tab_name]
        logger.info(f"Fetching albums from tab: {tab_name}")

        try:
            # Find header row
            header_row = self._find_header_row(worksheet)

            # Get column headers
            headers = []
            col_idx = 1
            while True:
                cell_value = worksheet.cell(row=header_row, column=col_idx).value
                if cell_value is None:
                    break
                headers.append(cell_value)
                col_idx += 1

            logger.debug(f"Tab '{tab_name}': Found {len(headers)} columns")

            # Create column index mapping
            col_mapping = {header: idx + 1 for idx, header in enumerate(headers)}

            # Validate required columns
            missing_columns = self.EXPECTED_COLUMNS - set(headers)
            if missing_columns:
                logger.warning(
                    f"Tab '{tab_name}' missing some expected columns: {missing_columns}"
                )

            # Parse data rows
            albums = []
            row_idx = header_row + 1
            while True:
                artist_cell = worksheet.cell(row=row_idx, column=col_mapping["Artist"])
                artist = artist_cell.value

                # Stop if we hit empty rows
                if not artist:
                    break

                album_cell = worksheet.cell(row=row_idx, column=col_mapping["Album"])
                album = album_cell.value

                # Skip rows without album name
                if not album:
                    row_idx += 1
                    continue

                # Extract Spotify URL
                spotify_cell = worksheet.cell(
                    row=row_idx, column=col_mapping["Spotify"]
                )
                spotify_url = self._extract_url_from_cell(spotify_cell)

                # Skip rows without Spotify URL
                if not spotify_url:
                    row_idx += 1
                    continue

                # Extract other fields
                # Note: Don't convert release_date to string - preserve datetime objects
                release_date_value = worksheet.cell(
                    row=row_idx, column=col_mapping.get("Release Date", 0)
                ).value
                normalized = {
                    "artist": str(artist).strip(),
                    "album": str(album).strip(),
                    "release_date": release_date_value,
                    "genre": str(
                        worksheet.cell(
                            row=row_idx, column=col_mapping.get("Genre / Subgenres", 0)
                        ).value
                        or ""
                    ).strip(),
                    "vocal_style": str(
                        worksheet.cell(
                            row=row_idx, column=col_mapping.get("Vocal Style", 0)
                        ).value
                        or ""
                    ).strip(),
                    "country": str(
                        worksheet.cell(
                            row=row_idx, column=col_mapping.get("Country / State", 0)
                        ).value
                        or ""
                    ).strip(),
                    "spotify_url": spotify_url,
                }

                # Add tab year if provided
                if tab_year is not None:
                    normalized["tab_year"] = tab_year

                albums.append(normalized)
                row_idx += 1

            logger.info(
                f"Tab '{tab_name}': Fetched {len(albums)} albums with Spotify URLs"
            )

            return albums

        except Exception as e:
            logger.error(f"Failed to parse tab '{tab_name}': {e}")
            raise

    def _extract_url_from_cell(self, cell) -> Optional[str]:
        """
        Extract URL from cell that might have hyperlink or HYPERLINK formula.

        Args:
            cell: openpyxl cell object

        Returns:
            URL string if found, None otherwise
        """
        # Check if cell has a hyperlink object
        if cell.hyperlink:
            return cell.hyperlink.target

        # Check if cell value is a HYPERLINK formula
        if (
            cell.value
            and isinstance(cell.value, str)
            and cell.value.startswith("=HYPERLINK")
        ):
            # Extract URL from =HYPERLINK("url", "text") formula
            match = re.search(r'=HYPERLINK\("([^"]+)"', cell.value)
            if match:
                return match.group(1)

        return None

    def _find_header_row(self, worksheet) -> int:
        """
        Find the header row in the worksheet (row with 'Artist' in first column).

        Args:
            worksheet: openpyxl worksheet object

        Returns:
            Row number (1-indexed) of the header row

        Raises:
            ValueError: If header row cannot be found
        """
        for row_idx in range(1, 20):  # Check first 20 rows
            first_cell = worksheet.cell(row=row_idx, column=1).value
            if first_cell == "Artist":
                logger.debug(f"Found header row at row {row_idx}")
                return row_idx

        raise ValueError("Could not find header row with 'Artist' column")

    def fetch_albums(self) -> List[Dict[str, str]]:
        """
        Fetch and parse album data from Google Sheets XLSX export.

        Returns:
            List of album dictionaries with normalized field names:
            - artist: Artist name
            - album: Album title
            - release_date: Release date string (e.g., "January 1")
            - genre: Genre/subgenre text
            - vocal_style: Vocal style text
            - country: Country or state
            - spotify_url: Full Spotify album URL

        Raises:
            requests.RequestException: If HTTP request fails
            ValueError: If required columns are missing or parsing fails
        """
        logger.info(f"Fetching XLSX from Google Sheets: {self.xlsx_url}")

        try:
            response = requests.get(self.xlsx_url, timeout=30)
            response.raise_for_status()

            # Load workbook from bytes
            xlsx_content = BytesIO(response.content)
            workbook = load_workbook(xlsx_content)
            worksheet = workbook.active

            # Extract year from active sheet name for release date parsing
            active_sheet_name = worksheet.title
            tab_year = extract_year(active_sheet_name)
            if tab_year:
                logger.debug(f"Extracted year {tab_year} from active sheet: {active_sheet_name}")

            # Find header row
            header_row = self._find_header_row(worksheet)

            # Get column headers
            headers = []
            col_idx = 1
            while True:
                cell_value = worksheet.cell(row=header_row, column=col_idx).value
                if cell_value is None:
                    break
                headers.append(cell_value)
                col_idx += 1

            logger.debug(f"Found {len(headers)} columns: {headers}")

            # Create column index mapping
            col_mapping = {header: idx + 1 for idx, header in enumerate(headers)}

            # Validate required columns
            missing_columns = self.EXPECTED_COLUMNS - set(headers)
            if missing_columns:
                logger.warning(
                    f"Sheet missing some expected columns: {missing_columns}. "
                    f"Available columns: {headers}"
                )

            # Parse data rows
            albums = []
            row_idx = header_row + 1
            while True:
                artist_cell = worksheet.cell(row=row_idx, column=col_mapping["Artist"])
                artist = artist_cell.value

                # Stop if we hit empty rows
                if not artist:
                    break

                album_cell = worksheet.cell(row=row_idx, column=col_mapping["Album"])
                album = album_cell.value

                # Skip rows without album name
                if not album:
                    row_idx += 1
                    continue

                # Extract Spotify URL
                spotify_cell = worksheet.cell(
                    row=row_idx, column=col_mapping["Spotify"]
                )
                spotify_url = self._extract_url_from_cell(spotify_cell)

                # Skip rows without Spotify URL
                if not spotify_url:
                    row_idx += 1
                    continue

                # Extract other fields
                # Note: Don't convert release_date to string - preserve datetime objects
                release_date_value = worksheet.cell(
                    row=row_idx, column=col_mapping.get("Release Date", 0)
                ).value
                normalized = {
                    "artist": str(artist).strip(),
                    "album": str(album).strip(),
                    "release_date": release_date_value,
                    "genre": str(
                        worksheet.cell(
                            row=row_idx, column=col_mapping.get("Genre / Subgenres", 0)
                        ).value
                        or ""
                    ).strip(),
                    "vocal_style": str(
                        worksheet.cell(
                            row=row_idx, column=col_mapping.get("Vocal Style", 0)
                        ).value
                        or ""
                    ).strip(),
                    "country": str(
                        worksheet.cell(
                            row=row_idx, column=col_mapping.get("Country / State", 0)
                        ).value
                        or ""
                    ).strip(),
                    "spotify_url": spotify_url,
                }

                # Add tab year if extracted from active sheet name
                if tab_year is not None:
                    normalized["tab_year"] = tab_year

                albums.append(normalized)
                row_idx += 1

            logger.info(
                f"Successfully fetched {len(albums)} albums with Spotify URLs "
                f"from Google Sheets"
            )

            return albums

        except requests.RequestException as e:
            logger.error(f"Failed to fetch XLSX from Google Sheets: {e}")
            raise

        except Exception as e:
            logger.error(f"Failed to parse XLSX data: {e}")
            raise

    def parse_release_date(self, date_value, year: int = None):
        """
        Parse release date from Google Sheets.

        The sheet may provide dates in different formats:
        - datetime object (openpyxl reads date cells as datetime)
        - String like "January 1", "May 15"
        - Month name only like "January"

        This method handles all formats and uses the tab year if provided.

        Args:
            date_value: Date value from sheet (datetime, str, or None)
            year: Year to use if not present in date_value (defaults to current year)

        Returns:
            datetime.date object, or None if parsing fails

        Note:
            Returns None for unparseable dates rather than raising exceptions
        """
        if not date_value:
            return None

        # If openpyxl returned a datetime object, use it directly
        if isinstance(date_value, datetime):
            # If year is provided and different from the datetime year, override it
            if year is not None and date_value.year != year:
                return date_value.replace(year=year).date()
            return date_value.date()

        # If it's already a date object
        if hasattr(date_value, 'year') and hasattr(date_value, 'month'):
            # If year is provided and different, override it
            if year is not None and date_value.year != year:
                return date_value.replace(year=year)
            return date_value

        # Handle string formats
        date_str = str(date_value).strip()
        if not date_str:
            return None

        # Use current year if not specified
        if year is None:
            year = datetime.now().year

        try:
            # Try parsing "Month Day" format (e.g., "January 1")
            date_str_with_year = f"{date_str}, {year}"
            parsed = datetime.strptime(date_str_with_year, "%B %d, %Y")
            return parsed.date()
        except ValueError:
            try:
                # Try "Month" only format (e.g., "January")
                date_str_with_year = f"{date_str} 1, {year}"
                parsed = datetime.strptime(date_str_with_year, "%B %d, %Y")
                return parsed.date()
            except ValueError:
                logger.warning(f"Could not parse release date: {date_value}")
                return None
